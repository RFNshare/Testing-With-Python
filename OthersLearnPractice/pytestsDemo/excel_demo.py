import openpyxl

book = openpyxl.load_workbook("PythonDemo.xlsx")
print(book)
sheet = book.active
print(sheet)
cell = sheet.cell(row=1, column=2)
print(cell.value)  # Read from excel

sheet.cell(row=4, column=1).value = "TestCase3"
print(book.save("PythonDemo.xlsx"))
# print(sheet.max_row)
# print(sheet.max_column)
#
# print(sheet['C3'].value)

# for row in range(1, sheet.max_row + 1):
#     if sheet.cell(row=row, column=1).value == "TestCase2":
#         for column in range(1, sheet.max_column + 1):
#             print(sheet.cell(row=row, column=column).value)


