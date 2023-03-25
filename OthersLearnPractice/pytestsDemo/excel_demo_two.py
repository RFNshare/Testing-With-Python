import openpyxl

book = openpyxl.load_workbook("PythonDemo.xlsx")
sheet = book.active
dic = {}
keys = []
values = [[] for i in range(1, sheet.max_column)]
k = 0
for i in range(1, 2):
    for j in range(1, sheet.max_column + 1):
        keys.append(sheet.cell(row=i, column=j).value)
for i in range(2, sheet.max_row + 1):
    for j in range(1, sheet.max_column + 1):
        values[k].append(sheet.cell(row=i, column=j).value)
    k += 1
lst_dic = []
for i in values:
    dic = dict(zip(keys, i))
    lst_dic.append(dic)
print(lst_dic)
