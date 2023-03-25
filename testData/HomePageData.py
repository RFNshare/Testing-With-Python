from pathlib import Path

import openpyxl


class HomePageData:
    test_homepage_data = [{"first_name": "Abdullah Al", "email": "aalfaroque@gmail.com", "password": "12345"},
                          {"first_name": "Faroque", "email": "rfnshare@gmail.com", "password": "4321"}]

    @staticmethod
    def get_test_data():
        EXCEL_PATH = Path(__file__).parent / 'PythonDemo.xlsx'
        book = openpyxl.load_workbook(EXCEL_PATH)
        sheet = book.active
        dic = {}
        keys = []
        values = [[] for i in range(1, sheet.max_column)]
        k = 0
        for i in range(1, 2):
            for j in range(1, sheet.max_column + 1):
                keys.append(sheet.cell(row=i, column=j).value)
        for i in range(2, sheet.max_row + 1):
            for j in range(1, sheet.max_column + 1):
                values[k].append(sheet.cell(row=i, column=j).value)
            k += 1
        lst_dic = []
        for i in values:
            dic = dict(zip(keys, i))
            if dic:
                lst_dic.append(dic)
        return lst_dic